from flask import Flask, render_template, request
from openpyxl import load_workbook

app = Flask(__name__)

@app.route("/")
def gallery():
    gal_list = []
    excel = load_workbook("Gallery.xlsx")
    page = excel["Лист1"]
    for row in page:
        picture = row[0].value
        description = row[1].value
        title = row[2].value
        lst = [title, description, picture]
        gal_list.append(lst)
    # gal_file = open("gal.txt", "r", encoding="utf-8")
    # gal_list = [row for row in gal_file]
    # gal_file.close()
    return render_template('gallery.html', gal_list=gal_list)

@app.route("/add")
def add():
    return render_template("add.html")

@app.route("/add-pic", methods=["POST"])
def add_pic():
    description = request.form.get("Description")
    picture = request.form.get("picture")
    title = request.form.get("title")
    
    excel = load_workbook("Gallery.xlsx")
    sheet = excel["Лист1"]
    sheet.append([picture, description, title])
    excel.save("Gallery.xlsx")
    return render_template("add.html")